"""GET /api/result/{id} 端点测试。sqlite PG + Redis 全 mock。"""
import pytest, httpx
from fastapi import FastAPI

from src.storage import query_results
from src.storage.pg_client import init_db
from src.web.routes.result import build_result_router


@pytest.fixture
async def client():
    await init_db("sqlite+aiosqlite:///:memory:")
    app = FastAPI()
    app.include_router(build_result_router())
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as c:
        yield c


async def _no_redis():
    """让 _get_redis 返回 None，强制走 PG。"""
    return None


@pytest.mark.asyncio
async def test_get_existing_result(client, monkeypatch):
    monkeypatch.setattr(query_results, "_get_redis", _no_redis)
    rid = await query_results.save_result("s", ["a"], [{"a": 1}])
    r = await client.get(f"/api/result/{rid}")
    assert r.status_code == 200
    body = r.json()
    assert body["columns"] == ["a"]
    assert body["rows"] == [{"a": 1}]
    assert body["total"] == 1


@pytest.mark.asyncio
async def test_get_missing_404(client, monkeypatch):
    monkeypatch.setattr(query_results, "_get_redis", _no_redis)
    r = await client.get("/api/result/nonexistent")
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_export_returns_xlsx(client, monkeypatch):
    """P4：/api/result/{id}/export 导出 Excel，表头+行正确。"""
    monkeypatch.setattr(query_results, "_get_redis", _no_redis)
    rid = await query_results.save_result("s", ["a", "b"],
                                          [{"a": 1, "b": "x"}, {"a": 2, "b": "y"}])
    r = await client.get(f"/api/result/{rid}/export")
    assert r.status_code == 200
    assert "spreadsheet" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    # 解析 xlsx 验证内容
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["a", "b"]   # 表头
    assert ws.max_row == 3                            # 表头 + 2 行数据


@pytest.mark.asyncio
async def test_export_missing_404(client, monkeypatch):
    monkeypatch.setattr(query_results, "_get_redis", _no_redis)
    r = await client.get("/api/result/nonexistent/export")
    assert r.status_code == 404
