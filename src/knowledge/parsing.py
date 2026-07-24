"""知识库上传文件解析：按扩展名转纯文本（供分段入库），统一页面导入入口。
- .txt/.md 直接解码
- .csv 按 meeting_insight 因果链格式转文本段落（字段不全则按普通表格行转）
- .docx 用 python-docx 提取段落（段落间空行保分段）
- .xlsx 用 openpyxl 按 sheet 转 markdown（标题 + 表格行）
依赖 python-docx / openpyxl（requirements 已列在 docs 转换脚本，页面导入复用同款解析）。"""
from __future__ import annotations

import csv
import io
from collections import defaultdict


def parse_to_text(filename: str, raw: bytes) -> str:
    """按扩展名解析上传文件 → 纯文本。未知扩展名按文本兜底解码。"""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        return _csv_to_text(raw)
    if ext == "docx":
        return _docx_to_text(raw)
    if ext == "xlsx":
        return _xlsx_to_text(raw)
    return raw.decode("utf-8", errors="ignore")   # txt/md/兜底


def _csv_to_text(raw: bytes) -> str:
    """meeting_insight CSV → 因果链文本段落（每条 insight 一段）。
    若无 root_cause/metric_category 字段（非 meeting_insight 格式），按普通表格行转文本。"""
    text = raw.decode("utf-8-sig", errors="ignore")   # utf-8-sig 去 BOM
    reader = csv.DictReader(io.StringIO(text))
    fields = reader.fieldnames or []
    if not any(k in fields for k in ("root_cause", "metric_category")):
        # 普通 CSV：每行 | 拼成一段
        rows = list(reader)
        return "\n\n".join(" | ".join(str(v) for v in r.values()) for r in rows) or text
    # meeting_insight 格式：每条 insight 一段（1 CSV = 1 文档，不按来源拆）
    sections: list[str] = []
    for r in reader:
        head = " ".join(x for x in [r.get("years"), r.get("operation_area") or r.get("ztype"),
                                     r.get("project"), r.get("metric_category")] if x)
        lines = [f"【会议洞察】{head or '会议洞察'}"]
        if r.get("metric_direction") or r.get("deviation_value"):
            lines.append(f"偏差：{r.get('metric_direction', '')} {r.get('deviation_value', '')}".strip())
        if r.get("root_cause"):
            lines.append(f"原因：{r['root_cause']}")
        if r.get("owner"):
            lines.append(f"责任人：{r['owner']}")
        if r.get("source_quote"):
            lines.append(f"原文：{r['source_quote'][:200]}")
        sections.append("\n".join(lines))
    return "\n\n".join(sections) if sections else text


def _docx_to_text(raw: bytes) -> str:
    """docx → 文本（段落间空行保分段）。"""
    from docx import Document
    paras = [p.text.strip() for p in Document(io.BytesIO(raw)).paragraphs if p.text.strip()]
    return "\n\n".join(paras)


def _xlsx_to_text(raw: bytes) -> str:
    """xlsx → 文本（每 sheet 一段：标题 + 表格行 | 拼保列结构）。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sections: list[str] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        filled = [[str(c).strip() for c in r if c is not None and str(c).strip()]
                  for r in ws.iter_rows(values_only=True)]
        filled = [r for r in filled if r]
        if filled:
            sections.append("\n".join([f"## {sn}"] + [" | ".join(r) for r in filled]))
    wb.close()
    return "\n\n".join(sections)
